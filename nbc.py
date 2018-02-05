#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sun Feb  4 12:59:56 2018

@author: Marine
"""

#from openpyxl import load_workbook
#
#
#wb = load_workbook(filename='/Users/Marine/Documents/1_KTH/4_Courses/7_Big_Data_in_Media_Technology/LABS/test_dataset .xlsx', read_only=True)
#ws = wb['big_data']
#
#for row in ws.rows:
#    for cell in row:
#        print(cell.value)
#        
        
import openpyxl

book = openpyxl.load_workbook('/Users/Marine/Documents/1_KTH/4_Courses/7_Big_Data_in_Media_Technology/LABS/test_dataset .xlsx')

sheet = book.active

#a1 = sheet['A1']
#a2 = sheet['A2']
#a3 = sheet.cell(row=3, column=1)

#print(a1.value)
#print(a2.value) 
#print(a3.value)

rows = sheet.rows

#for row in rows:
#    for cell in row:
#        print(cell.value)

#from pandas import DataFrame, read_excel

#import matplotlib.pyplot as plt
import pandas as pd

# Returns a DataFrame
df_test = pd.read_excel('/Users/Marine/Documents/1_KTH/4_Courses/7_Big_Data_in_Media_Technology/LABS/test_dataset .xlsx', sheet_name='Sheet1', header=None, names=['Comments'])
print(df_test.dtypes)
print(df_test.Comments.dtypes)

df = pd.read_csv('/Users/Marine/Documents/1_KTH/4_Courses/7_Big_Data_in_Media_Technology/LABS/training_dataset.txt', sep="	", header=None, names=["target", "data"])
print(df.dtypes)
print(df.data.dtypes)
print(df.target.dtypes)

df.target_names = ['negative','positive']



from nltk.classify.scikitlearn import SklearnClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.naive_bayes import MultinomialNB

classif = SklearnClassifier(GaussianNB())

#Tokenizing

from sklearn.feature_extraction.text import CountVectorizer
count_vect = CountVectorizer()
X_train_counts = count_vect.fit_transform(df.data)
X_train_counts.shape


#From occurances to frequencies
from sklearn.feature_extraction.text import TfidfTransformer
tf_transformer = TfidfTransformer(use_idf=False).fit(X_train_counts)
X_train_tf = tf_transformer.transform(X_train_counts)
X_train_tf.shape

tfidf_transformer = TfidfTransformer()
X_train_tfidf = tfidf_transformer.fit_transform(X_train_counts)
X_train_tfidf.shape

#classif.fit(X_train_tfidf, df.target)
clf = MultinomialNB().fit(X_train_tfidf, df.target)
classif.train(df)


docs_new = ['I loved it !', 'breakfast was amazing', 'I hated it', 'I dont know what to think of it']
X_new_counts = count_vect.transform(docs_new)
X_new_tfidf = tfidf_transformer.transform(X_new_counts)

predicted = clf.predict(X_new_tfidf)

for doc, category in zip(docs_new, predicted):
    print('%r => %s' % (doc, df.target_names[category]))


#classif.train(df)
#print("GaussianNB accuracy percent:",nltk.classify.accuracy(GaussianNB, testing_set))


#from sklearn import datasets
#iris1 = df
#iris = datasets.load_iris()
#from sklearn.naive_bayes import GaussianNB
#gnb = GaussianNB()
#y_pred = gnb.fit(iris.data, iris.target).predict(iris.data)
#print("Number of mislabeled points out of a total %d points : %d"
#      % (iris.data.shape[0],(iris.target != y_pred).sum()))